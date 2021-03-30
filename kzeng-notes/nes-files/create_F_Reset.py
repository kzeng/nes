# --------------------------------------------------------------------------------
# 生成F_Reset.asm 中的代码 片段
# 输入#1：游戏库EXCEL文件, 如Nes-Games-List.xlsx
# 输入#2：.IDX 文件， 如: Romfile.idx

# 生成以下信息
# A_NameTbl:
#                 DW      Name_001
#                 DW      Name_002
#                 DW      Name_003
# ...

# 和

# Name_001:
#         DB      'CONTRA JPN 30P',$00
# Name_002:
#         DB      'SUPER CONTRA 30P',$00
# Name_003:
#         DB      'CONTRA FORCE',$00
# ...

# 和
# A_GameTbl:
# 			dw	Game_001
# 			dw	Game_002
# 			dw	Game_003
# ...

# 和

# ;===============================================================================
# ;魂斗罗1代	日文版：预置$07FF，D6-D4选关，D3允许30条命，D2-D0选武器
# ;-------------------------------------------------------------------------------
# Game_001:
# 		SetupGame	$0040000,P08,$0060000,15,0,1
# 			dw	Game001_Set
# --------------------------------------------------------------------------------

import os
import time
from openpyxl import load_workbook,Workbook


F_Reset_Section = "./F_Reset_Section.asm"
source_file = './Nes-Games-List.xlsx'
idx_file = "./Romfile.idx"

idx_dict = {}

wb = load_workbook(source_file)
ws1 = wb["Sheet1"]

f = open(F_Reset_Section, 'w')
f.write("{} \n".format('A_NameTbl:'))
for row in ws1.iter_rows(min_row=2):
    f.write("                {}      {}\n".format('DW', 'Name_' +  str(row[0].value).zfill(3) ))

game_file_name=""
f.write("{} \n".format(';-------------------------------------------------------------------------------'))
for row in ws1.iter_rows(min_row=2):
    f.write("{}\n".format('Name_' +  str(row[0].value).zfill(3) ))
    for gn in row[4].value.split('.')[0].split('_'):
        game_file_name += gn.upper() + " "
    f.write("        {}      '{}',{}\n".format('DB', game_file_name.strip(), '$00' ))
    game_file_name=""
    
f.write("{} \n".format(';==============================================================================='))
f.write("{} \n".format('A_GameTbl:'))
for row in ws1.iter_rows(min_row=2):
    f.write("\t\t\t{}\t{}\n".format('dw', 'Game_' + str(row[0].value).zfill(3) ))

    
# 例如，Game_001下面一行语句，表示魂斗罗1代的程序数据起点为0x40000，
# 程序数据长度为P08，图像数据起点为0x60000，图像数据长度为16，纵向镜像，需要预设。
# ;===============================================================================
# ;魂斗罗1代	日文版：预置$07FF，D6-D4选关，D3允许30条命，D2-D0选武器
# ;-------------------------------------------------------------------------------
# Game_001:
# 		SetupGame	$0040000,P08,$0060000,15,0,1
# 			dw	Game001_Set
f1 = open(idx_file, 'r')
for item in f1:
#     print(item)
    if item.split('\t')[0].strip() == 'F_Reset.bin':
        print('Skip F_Reset.bin file...')
        continue
    else:
        idx_dict.update({item.split('\t')[0].strip() : item.split('\t')[2].strip() })
        
for row in ws1.iter_rows(min_row=2):
    f.write("{} \n".format(';==============================================================================='))
    f.write("; 文件名:{}\t 中文名:{}\t 版本:{}\t 有无预置:{} \n".format(row[4].value, '???', '???', '???'))
    f.write("{} \n".format(';-------------------------------------------------------------------------------'))
    f.write("{}: \n".format('Game_' + str(row[0].value).zfill(3) ))        
    chr_addr = '{:07x}'.format(int('0x'+idx_dict[row[4].value], 16) + int(row[5].value)*16*1024).upper()
    f.write("\t\tSetupGame\t{},{},{},{},{},{}\n".format( '$'+ idx_dict[row[4].value], 'P'+str(row[5].value).zfill(2), '$'+chr_addr, row[6].value, row[7].value, 0 ))
    #f.write("\t\t\tdw\tGame{}_Set\n".format( str(row[0].value).zfill(3) ))


f.close()
print('生成 F_Reset.asm中的代码片段：' + F_Reset_Section + ' Ok!')


print('以下为F_Reset.asm中的代码片段：')
# show created F_Reset_Section content
f2 = open(F_Reset_Section, 'r')
for item in f2:
    print(item, end='')
f2.close()

