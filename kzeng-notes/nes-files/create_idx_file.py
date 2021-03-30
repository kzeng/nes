## 生成 .IDX 文件优化版本
## 文件名: Romfile.idx
# 文件格式如下：
# AIsland4.nes 	 0000010 	 0000000 
# AIsland1.nes 	 0000010 	 0180000 
# Dr_Mario.nes 	 0000010 	 02D0000 
# F_Reset.bin 	 0000000 	 007E000 
# BattKid2.nes 	 0000010 	 0080038 

# 输入： 游戏列表的电子表格文件，如：Nes-Game-List.xlsx
# 输出： IDX 文件, 如：Testnes.idx

import os
import time
from openpyxl import load_workbook,Workbook

source_file = './Nes-Games-List.xlsx'
idx_file = "./Romfile.idx"

F_reset_addr = int(0x007E000/1024)
little_file_start_addr = '0000010'
rom_size_sum = 0
the_first_item = True
nes_files_dict = {}
pick_nes_files = {}

f = open(idx_file, 'w')

wb = load_workbook(source_file)
ws1 = wb["Sheet1"]

for row in ws1.iter_rows(min_row=2):
    rom_size = int(row[5].value) * 16 * 1024 + int(row[6].value) * 8 * 1024
    nes_files_dict.update( {row[4].value: int(rom_size/1024) } )
    nes_files_dict_sorted = sorted(nes_files_dict.items(), key=lambda nes_files_dict: nes_files_dict[1], reverse=True)

# nes_files_dict_sorted
# 挑选出ROMFILE 前段空间安放的nes文件，原则尽量占满前端空间(0x007E000 即504K)
for p in nes_files_dict_sorted:
    if p[1] == F_reset_addr:
        pick_nes_files.update({p[0] : p[1]})
        break
    elif p[1] < F_reset_addr:
        pick_nes_files.update({p[0] : p[1]})
        F_reset_addr = F_reset_addr - p[1]
        continue

print('以下挑选出来的nes 文件安放在F_reset.bin 之前段')
print(pick_nes_files)
    
# 生成Testnes.idx (即Romfile.idx)。 分三段，即F_reset.bin 之前一段，F_reset.bin 自身一段， F_reset.bin之后一段
print("开始安放nes文件到 .IDX 文件...")
# F_reset.bin 之前一段
for nes_file in pick_nes_files:
    if the_first_item:
        large_file_local_addr = rom_size_sum
        the_first_item = False
    else:
        rom_size = int(pick_nes_files[nes_file]) * 16 * 1024 + int(pick_nes_files[nes_file]) * 8 * 1024
        rom_size_sum += rom_size
      
    f.write("{} \t {} \t {} \n".format(nes_file, little_file_start_addr, '{:07x}'.format(rom_size_sum).upper() ))
    
# F_reset.bin 自身一段
f.write("F_Reset.bin \t 0000000 \t 007E000 \n")

# F_reset.bin之后一段
rom_size_sum1 = F_reset_addr + 8 #F_reset.bin size 8k ???

for p in nes_files_dict_sorted:
    if p[0] in pick_nes_files:
        #print('This nes file in first section, skip!')
        continue
    else:
        rom_size = p[1] * 1024
        rom_size_sum1 += rom_size
        
    f.write("{} \t {} \t {} \n".format(p[0], little_file_start_addr, '{:07x}'.format(rom_size_sum1).upper() ))

f.close()

# show created .idx content
f1 = open(idx_file, 'r')
for item in f1:
    print(item, end='')
f1.close()

print('生成 ' + idx_file + ' Ok!')
