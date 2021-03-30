# 分析 .NES 文件格式
# 输入： nes 格式文件名
# 输出: 包含NES头各字段信息的字典类型数据
import struct
import os
import time
from openpyxl import load_workbook,Workbook

def nes_parser(filename='Spartanx.nes', offset=0x0, size=1):
    # print('parsing nes file (' + filename + ')...')
    with open(filename, 'rb') as f:
        data = f.read()

    if size == 1:
        format = 'B'
    elif size == 2:
        format = '<H'
    elif size == 4:
        format = '<I'
    elif size == 8:
        format = '<Q'
    else:
        print('invalid size', size)

    value = struct.unpack(format, data[offset : offset + size])[0]
    # print('value=', hex(value), value)
    # return hex(value)
    return value

# nes_parser(offset=0x1)

def nes_summary(filename='Spartanx.nes'):
    nes_summary = {}
    file_flag = chr(nes_parser(filename, offset=0x0)) + chr(nes_parser(filename, offset=0x1)) + chr(nes_parser(filename, offset=0x2))
    num_16k_rom = nes_parser(filename, offset=0x4)
    num_8k_vrom = nes_parser(filename, offset=0x5)
    
    nes_summary["filename"] = filename
    nes_summary["file_flag"] = file_flag
    nes_summary["num_16k_rom"] = num_16k_rom
    nes_summary["num_8k_vrom"] = num_8k_vrom
    
    # ROM Control Byte#1
    crb1 = '{:08b}'.format(nes_parser(filename, offset=0x6))    
    # '01000001'
    # D0：1＝垂直镜像，0＝水平镜像  
    nes_summary["mirroring"] = crb1[-1]
    # D1：1＝有电池记忆，SRAM地址 6000− 7FFF
    nes_summary["battery"] = crb1[-2]
    # D2：1＝在 7000− 71FF有一个512字节的trainer D3：1＝4屏幕VRAM布局
    nes_summary["trainer"] = crb1[-3]
    # D3：1＝4屏幕VRAM布局
    nes_summary["screen"] = crb1[-4]
    # D4－D7：ROM Mapper的低4位
    nes_summary["mapper_lower_4bits"] = '{}{}{}{}'.format(crb1[-5], crb1[-6], crb1[-7], crb1[-8])
    
    # ROM Control Byte#2
    crb2 = '{:08b}'.format(nes_parser(filename, offset=0x7))
    # D4－D7：ROM Mapper的高4位
    nes_summary["mapper_upper_4bits"] = '{}{}{}{}'.format(crb2[-5], crb2[-6], crb2[-7], crb2[-8])    
#     print(nes_summary)
    return(nes_summary)



# 自动生成有列表电子表格文件
# 这个电子表格文件将作为自动生成 F_Reset.asm, Testnes.idx 代码内容的！！重要！！输入源
# 输入： 读取./128in1 目录下的所有.nes文件， 按.NES 格式分析文件头部，将必要信息提取填入到电子表格中 
# 输出：包含 序号	版本	文件名	 程序容量(PRG)	图像容量(CHR)	镜像 等字段的表格，如： Nes-Games-List.XLSX
#

nes_files_dir = './128in1/'
# nes_games_list = 'Nes-Games-List-' + str(time.time()) + '.xlsx'
nes_games_list = 'Nes-Games-List' + '.xlsx'

print('Start create ' + nes_games_list + '...')

wb = Workbook()
ws = wb.create_sheet("Sheet1")

ws.append(['序号', '版本', '英文名',  '中文名', '文件名', '程序容量(PRG)16KxM', '图像容量(CHR)8KxN','镜像(1垂直 0水平)', '预置'])

idx = 1
for root, path, files_name in os.walk(nes_files_dir):
    #print(files_name)
    for nes in files_name:
        ret = nes_summary(root + nes)
        ret_list = [idx, '-', '-', '-', nes, ret['num_16k_rom'], ret['num_8k_vrom'], ret['mirroring'], '-']
        ws.append(ret_list)
        idx = idx + 1

wb.save(nes_games_list)
print('Create ' + nes_games_list + ' Ok!')
